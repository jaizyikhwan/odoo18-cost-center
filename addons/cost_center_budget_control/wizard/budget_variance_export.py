# -*- coding: utf-8 -*-
"""Wizard to export Budget Variance report as CSV or XLSX.

A simple, framework-agnostic exporter. CSV uses Python's standard ``csv``
module (no external dependencies). XLSX uses ``openpyxl`` when available,
falling back to CSV with a notice when the library is missing.

The wizard is launched from the Budget Plan list/form view and can be
filtered by:
  - A set of budget plans (selected from the list)
  - Date range
  - Cost center
  - State

The output file is streamed back via ``web.Binary`` so the browser gets
a direct download.
"""
import base64
import csv
import io
import logging
from datetime import datetime

from odoo import api, fields, models, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class BudgetVarianceExport(models.TransientModel):
    _name = "budget.variance.export"
    _description = "Export Budget Variance Report"

    plan_ids = fields.Many2many(
        "budget.plan",
        string="Budget Plans",
        help="Leave empty to use the filter criteria below.",
    )
    date_from = fields.Date(string="Start Date")
    date_to = fields.Date(string="End Date")
    cost_center_id = fields.Many2one(
        "cost.center",
        string="Cost Center",
    )
    state_filter = fields.Selection([
        ("all", "All"),
        ("approved", "Approved Only"),
        ("active", "Currently Active"),
        ("over_budget", "Over Budget"),
    ], string="Filter", default="all", required=True)
    format = fields.Selection([
        ("csv", "CSV"),
        ("xlsx", "XLSX (Excel)"),
    ], string="Output Format", default="xlsx", required=True)
    include_company_currency = fields.Boolean(
        string="Include Company Currency Column",
        default=True,
        help="Add a column showing the amount in the company's reporting currency. "
             "Only useful for multi-currency budgets.",
    )

    file_name = fields.Char(string="File Name", readonly=True)
    file_data = fields.Binary(string="File", readonly=True)

    # -------------------------------------------------------------------------
    # DEFAULTS
    # -------------------------------------------------------------------------

    @api.model
    def default_get(self, fields_list):
        res = super().default_get(fields_list)
        # Pre-select plans from the list view
        if self.env.context.get("active_model") == "budget.plan" and self.env.context.get("active_ids"):
            res["plan_ids"] = [(6, 0, self.env.context["active_ids"])]
        return res

    # -------------------------------------------------------------------------
    # CORE
    # -------------------------------------------------------------------------

    def _resolve_plan_ids(self):
        """Resolve which plans to export based on the criteria."""
        self.ensure_one()
        if self.plan_ids:
            return self.plan_ids
        domain = [("company_id", "=", self.env.company.id)]
        if self.date_from:
            domain.append(("date_to", ">=", self.date_from))
        if self.date_to:
            domain.append(("date_from", "<=", self.date_to))
        if self.cost_center_id:
            domain.append(("cost_center_id", "=", self.cost_center_id.id))
        if self.state_filter == "approved":
            domain.append(("state", "=", "approved"))
        elif self.state_filter == "active":
            today = fields.Date.today()
            domain += [
                ("state", "=", "approved"),
                ("date_from", "<=", today),
                ("date_to", ">=", today),
            ]
        elif self.state_filter == "over_budget":
            domain.append(("line_ids.available_amount", "<", 0))
        return self.env["budget.plan"].search(domain, order="date_from desc, name")

    def _collect_rows(self, plans):
        """Return a list of dicts, one per budget line."""
        rows = []
        for plan in plans:
            for line in plan.line_ids:
                rows.append({
                    "plan": plan.name,
                    "state": plan.state,
                    "cost_center": plan.cost_center_id.complete_name,
                    "company": plan.company_id.name,
                    "currency": plan.currency_id.name,
                    "company_currency": plan.company_currency_id.name,
                    "is_multi_currency": "Y" if plan.is_multi_currency else "N",
                    "account": f"{line.account_id.code} {line.account_id.name}",
                    "planned": line.planned_amount,
                    "actual": line.actual_amount,
                    "po_committed": line.po_committed_amount,
                    "committed": line.committed_amount,
                    "available": line.available_amount,
                    "remaining": line.remaining_amount,
                    "usage_percent": line.usage_percent,
                    "alert_level": line.alert_level,
                    "date_from": plan.date_from,
                    "date_to": plan.date_to,
                    "company_currency_planned": line.planned_amount_company_currency,
                    "company_currency_actual": line.actual_amount_company_currency,
                    "company_currency_committed": line.committed_amount_company_currency,
                    "company_currency_available": line.available_amount_company_currency,
                })
        return rows

    def _headers(self):
        base = [
            "Budget Plan", "State", "Cost Center", "Company",
            "Currency", "Company Currency", "Multi-Currency",
            "Account",
            "Planned", "Actual", "PO Committed", "Committed", "Available",
            "Remaining", "Usage %", "Alert Level",
            "Date From", "Date To",
        ]
        if self.include_company_currency:
            base += [
                "Planned (Co.Curr)", "Actual (Co.Curr)",
                "Committed (Co.Curr)", "Available (Co.Curr)",
            ]
        return base

    # -------------------------------------------------------------------------
    # FORMAT-SPECIFIC BUILDERS
    # -------------------------------------------------------------------------

    def _build_csv(self, rows):
        buf = io.StringIO()
        writer = csv.writer(buf, quoting=csv.QUOTE_MINIMAL)
        writer.writerow(self._headers())
        for r in rows:
            row = [
                r["plan"], r["state"], r["cost_center"], r["company"],
                r["currency"], r["company_currency"], r["is_multi_currency"],
                r["account"],
                f"{r['planned']:.2f}", f"{r['actual']:.2f}",
                f"{r['po_committed']:.2f}", f"{r['committed']:.2f}",
                f"{r['available']:.2f}", f"{r['remaining']:.2f}",
                f"{r['usage_percent']:.2f}", r["alert_level"],
                r["date_from"], r["date_to"],
            ]
            if self.include_company_currency:
                row += [
                    f"{r['company_currency_planned']:.2f}",
                    f"{r['company_currency_actual']:.2f}",
                    f"{r['company_currency_committed']:.2f}",
                    f"{r['company_currency_available']:.2f}",
                ]
            writer.writerow(row)
        return buf.getvalue().encode("utf-8")

    def _build_xlsx(self, rows):
        if not OPENPYXL_AVAILABLE:
            raise UserError(_(
                "The 'openpyxl' Python library is not installed on the server. "
                "Please install it (pip install openpyxl) or use the CSV format."
            ))
        wb = Workbook()
        ws = wb.active
        ws.title = "Budget Variance"

        header_fill = PatternFill(start_color="714B67", end_color="714B67", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        center = Alignment(horizontal="center", vertical="center")

        headers = self._headers()
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        money_cols_planned = [9, 10, 11, 12, 13, 14]
        if self.include_company_currency:
            money_cols_planned += [19, 20, 21, 22]
        number_cols = [15]

        for row_idx, r in enumerate(rows, start=2):
            data = [
                r["plan"], r["state"], r["cost_center"], r["company"],
                r["currency"], r["company_currency"], r["is_multi_currency"],
                r["account"],
                r["planned"], r["actual"], r["po_committed"], r["committed"],
                r["available"], r["variance"], r["usage_percent"], r["alert_level"],
                str(r["date_from"]), str(r["date_to"]),
            ]
            if self.include_company_currency:
                data += [
                    r["company_currency_planned"], r["company_currency_actual"],
                    r["company_currency_committed"], r["company_currency_available"],
                ]
            for col_idx, val in enumerate(data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=val)
            for c in money_cols_planned:
                ws.cell(row=row_idx, column=c).number_format = "#,##0.00"
            for c in number_cols:
                ws.cell(row=row_idx, column=c).number_format = "0.00"

        for col_letter, _ in enumerate(headers, start=1):
            ws.column_dimensions[chr(64 + col_letter) if col_letter <= 26 else "A" + chr(64 + col_letter - 26)].width = 18

        ws.freeze_panes = "A2"
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # -------------------------------------------------------------------------
    # ACTIONS
    # -------------------------------------------------------------------------

    def action_export(self):
        self.ensure_one()
        plans = self._resolve_plan_ids()
        if not plans:
            raise UserError(_("No budget plans match the selected criteria."))

        rows = self._collect_rows(plans)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if self.format == "xlsx":
            data = self._build_xlsx(rows)
            ext = "xlsx"
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        else:
            data = self._build_csv(rows)
            ext = "csv"
            content_type = "text/csv"

        file_name = f"budget_variance_{timestamp}.{ext}"
        self.write({
            "file_name": file_name,
            "file_data": base64.b64encode(data),
        })

        return {
            "type": "ir.actions.act_url",
            "url": f"/web/content/?model=budget.variance.export&id={self.id}"
                   f"&filename_field=file_name&field=file_data&download=true",
            "target": "self",
        }
